#!/usr/bin/env python3
"""
EXTRACTOR DEL UNIVERSO DE PROYECTOS
===================================

Lee la pestaña "Master Inmuebles Pro" del máster y produce universo.json,
que es lo que consume motor.js.

  python3 extraer.py Master_Inmuebles_Reental.xlsx > universo.json

MAPEO DE COLUMNAS
-----------------
Deducido de las fórmulas de la plantilla y verificado contra las cabeceras
reales del máster (fila 3). Los VLOOKUP de la plantilla usan índices
relativos al rango B:CK, así que el índice N corresponde a la columna N+1.

  A   ID                          O   Ubicación
  B   Nombre del proyecto         P   Tipología de explotación
  C   ESTADO                      Q   Tipología de Dividendo
  K   Px Emisión Token            L   Divisa
  BW  Descripción                 CQ  Link a Dossier Comercial
  CR  Link a Whitepaper (ES)

RENTABILIDADES — la plantilla elige entre estimada y real según el estado:

  Rendimientos recurrentes anualizados
    real      BB / CF / CJ   (Reentel / RP / SR)  si CERRADO o EN EXPLOTACIÓN
    estimada  X  / AB / AF                        en cualquier otro caso

  Plusvalía
    real      BI / BN / BS                        solo si CERRADO
    estimada  Y  / AC / AG                        en cualquier otro caso

  El motivo: un proyecto en explotación ya paga rentas reales, pero al no
  haberse vendido todavía no tiene plusvalía real.

FECHA DE FIN
------------
La columna AS ("Estimación Nº Meses pendientes") NO es fiable: devuelve 0 o
#NUM! en los proyectos que han sobrepasado su fecha de fin y a los que no se
les ha rellenado la columna auxiliar CA. Por eso los meses restantes se
calculan aquí desde la cascada que ya usa la plantilla para la fecha de fin:

  CA  Fecha fin Estimada Auxiliar (para los que la han sobrepasado)
  BD  Real fecha de fin
  I   Estimación fecha fin desde Financiación
  F   Estimación fecha de fin desde Lanzamiento

Se toma la primera que tenga valor.
"""

import sys
import json
import datetime
from openpyxl import load_workbook

HOJA = 'Master Inmuebles Pro'
FILA_DATOS = 4
ESTATUS = ['Reentel', 'ReentelPro', 'SuperReentel']

COL = {
    'id': 1, 'nombre': 2, 'estado': 3,
    'pxEmision': 11, 'divisa': 12, 'ubicacion': 15,
    'tipExpl': 16, 'tipDiv': 17,
    'mesesAS': 45, 'inicioRenta': 5, 'descripcion': 75, 'dossier': 95, 'whitepaper': 96,
}
FECHA_FIN_CASCADA = [79, 56, 9, 6]          # CA, BD, I, F
NOMBRE_CASCADA = {79: 'CA', 56: 'BD', 9: 'I', 6: 'F'}

EST_ALQ = {'Reentel': 24, 'ReentelPro': 28, 'SuperReentel': 32}   # X  AB AF
EST_PLUS = {'Reentel': 25, 'ReentelPro': 29, 'SuperReentel': 33}  # Y  AC AG
REAL_ALQ = {'Reentel': 54, 'ReentelPro': 84, 'SuperReentel': 88}  # BB CF CJ
REAL_PLUS = {'Reentel': 61, 'ReentelPro': 66, 'SuperReentel': 71} # BI BN BS


def numero(v):
    """Devuelve el valor solo si es numérico. Las celdas con #NUM! o texto
    llegan como str y se descartan."""
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def texto(v):
    return str(v).strip() if v is not None else ''


def meses_entre(desde, hasta):
    """Meses enteros, igual que DATEDIF(...,"m"). Puede ser negativo si el
    proyecto ya ha sobrepasado su fecha de fin."""
    m = (hasta.year - desde.year) * 12 + (hasta.month - desde.month)
    if hasta.day < desde.day:
        m -= 1
    return m


def extraer(ruta, hoy=None):
    hoy = hoy or datetime.date.today()
    ws = load_workbook(ruta, data_only=True)[HOJA]
    universo = []

    for r in range(FILA_DATOS, ws.max_row + 1):
        pid = texto(ws.cell(r, COL['id']).value)
        if not pid:
            continue

        estado = texto(ws.cell(r, COL['estado']).value).upper()
        usa_real_alq = estado in ('CERRADO', 'EN EXPLOTACIÓN')
        usa_real_plus = estado == 'CERRADO'

        fin, via = None, '-'
        for c in FECHA_FIN_CASCADA:
            v = ws.cell(r, c).value
            if hasattr(v, 'date'):
                fin, via = v.date(), NOMBRE_CASCADA[c]
                break

        registro = {
            'id': pid,
            'nombre': texto(ws.cell(r, COL['nombre']).value),
            'estado': estado,
            'ubicacion': texto(ws.cell(r, COL['ubicacion']).value),
            'tipologiaExplotacion': texto(ws.cell(r, COL['tipExpl']).value),
            'tipologiaDividendos': texto(ws.cell(r, COL['tipDiv']).value),
            'pxEmision': numero(ws.cell(r, COL['pxEmision']).value),
            'divisa': texto(ws.cell(r, COL['divisa']).value),
            'fechaFin': fin.isoformat() if fin else None,
            'inicioRenta': (lambda v: v.date().isoformat() if hasattr(v,'date') else None)(ws.cell(r, COL['inicioRenta']).value),
            'fuenteFechaFin': via,
            'mesesRestantes': meses_entre(hoy, fin) if fin else None,
            'mesesAS': numero(ws.cell(r, COL['mesesAS']).value),
            'descripcion': texto(ws.cell(r, COL['descripcion']).value),
            'dossier': texto(ws.cell(r, COL['dossier']).value),
            'whitepaper': texto(ws.cell(r, COL['whitepaper']).value),
            'rentAlquiler': {},
            'plusvalia': {},
            'fuenteRentabilidad': ('real' if usa_real_alq else 'estimada')
                                  + '/' + ('real' if usa_real_plus else 'estimada'),
        }

        for e in ESTATUS:
            col_a = REAL_ALQ[e] if usa_real_alq else EST_ALQ[e]
            col_p = REAL_PLUS[e] if usa_real_plus else EST_PLUS[e]
            registro['rentAlquiler'][e] = numero(ws.cell(r, col_a).value) or 0
            registro['plusvalia'][e] = numero(ws.cell(r, col_p).value) or 0

        universo.append(registro)

    return universo


def avisos(universo):
    """Problemas de calidad de dato que conviene revisar en el máster."""
    out = []
    for p in universo:
        if p['estado'] in ('CERRADO', 'NO LANZADO'):
            continue
        if p['mesesRestantes'] is None:
            out.append(f"{p['id']}: sin fecha de fin en ninguna de las 4 columnas")
        elif p['mesesRestantes'] < 0:
            out.append(
                f"{p['id']}: fecha de fin superada el {p['fechaFin']} "
                f"({abs(p['mesesRestantes'])} meses) y sin CA rellenada"
            )
        if not p['pxEmision']:
            out.append(f"{p['id']}: sin precio de emisión")
        if not p['ubicacion']:
            out.append(f"{p['id']}: sin ubicación")
    return out


if __name__ == '__main__':
    ruta = sys.argv[1] if len(sys.argv) > 1 else 'Master_Inmuebles_Reental.xlsx'
    universo = extraer(ruta)
    for a in avisos(universo):
        print('AVISO ' + a, file=sys.stderr)
    print(f"{len(universo)} proyectos extraídos", file=sys.stderr)
    json.dump(universo, sys.stdout, ensure_ascii=False, indent=1)
